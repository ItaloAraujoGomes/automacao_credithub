import os
import pandas as pd
import win32com.client
import re
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE_DIRECTORY = r"C:\Users\7981\Desktop\CreditHub"
FILE_NAME = "dados_empresas_atualizados.xlsx"
FILE_PATH = os.path.join(BASE_DIRECTORY, FILE_NAME)

# Funções auxiliares
def clean_razao_social(raw_name):
    # Remove palavras como "CPF", ou outras que não façam parte da razão social
    cleaned_name = re.sub(r"CPF|CNPJ|CNPJ da empresa|[0-9]{2,}", "", raw_name.strip())
    return cleaned_name.strip()

def format_cnpj(cnpj):
    return f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"

def process_emails(subject_filters, email_type):
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)

    data = []
    time_threshold = datetime.now() - timedelta(days=7)

    for item in inbox.Items:
        if item.Class == 43:
            try:
                received_time = item.ReceivedTime.replace(tzinfo=None)
                if received_time >= time_threshold:
                    if any(re.search(rf"{filter}", item.Subject, re.IGNORECASE) for filter in subject_filters):
                        body = item.Body
                        razao_social = re.search(r"Nome/Razão Social:\s*([\w\s\.\-]+)", body)
                        cpf_cnpj = re.search(r"(\d{14})", body)

                        if razao_social and cpf_cnpj:
                            data.append({
                                "Razão Social": clean_razao_social(razao_social.group(1)),
                                "CPF/CNPJ": cpf_cnpj.group(1).strip(),
                                "Tipo": email_type
                            })
            except Exception as e:
                print(f"Erro ao processar e-mail com assunto '{item.Subject}': {e}")
    return data

def update_excel(data):
    log_changes = []
    if os.path.exists(FILE_PATH):
        df_existing = pd.read_excel(FILE_PATH)
    else:
        df_existing = pd.DataFrame(columns=["Razão Social", "CPF/CNPJ", "Protestos", "Processos", "CCF", "Última Modificação"])

    for entry in data:
        cnpj = entry["CPF/CNPJ"]
        tipo = entry["Tipo"]
        today = datetime.now().strftime("%d/%m/%Y")
        existing_row = df_existing[df_existing["CPF/CNPJ"] == cnpj]

        if not existing_row.empty:
            idx = existing_row.index[0]
            old_data = existing_row.iloc[0].to_dict()
            if tipo == "Protestos":
                df_existing.at[idx, "Protestos"] += 1
            elif tipo == "Processos":
                df_existing.at[idx, "Processos"] += 1
            elif tipo == "CCF":
                df_existing.at[idx, "CCF"] += 1
            df_existing.at[idx, "Última Modificação"] = today
            new_data = df_existing.iloc[idx].to_dict()
            log_changes.append({
                "Razão Social": old_data["Razão Social"],
                "CPF/CNPJ": old_data["CPF/CNPJ"],
                "Campo Alterado": tipo,
                "Antes": old_data[tipo],
                "Depois": new_data[tipo],
                "Data da Alteração": today
            })
        else:
            new_row = {
                "Razão Social": entry["Razão Social"],
                "CPF/CNPJ": cnpj,
                "Protestos": 1 if tipo == "Protestos" else 0,
                "Processos": 1 if tipo == "Processos" else 0,
                "CCF": 1 if tipo == "CCF" else 0,
                "Última Modificação": today
            }
            df_existing = pd.concat([df_existing, pd.DataFrame([new_row])], ignore_index=True)
            log_changes.append({
                "Razão Social": entry["Razão Social"],
                "CPF/CNPJ": cnpj,
                "Campo Alterado": "Novo Registro",
                "Antes": "-",
                "Depois": "Adicionado",
                "Data da Alteração": today
            })

    df_existing["CPF/CNPJ"] = df_existing["CPF/CNPJ"].apply(format_cnpj)
    df_existing.to_excel(FILE_PATH, index=False)

    wb = openpyxl.load_workbook(FILE_PATH)
    ws = wb.active

    # Estilização da planilha
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                    top=Side(border_style="thin"), bottom=Side(border_style="thin"))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")  # Centralizando o conteúdo das células

    # Ajuste de largura das colunas
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                column_widths[cell.column_letter] = max((column_widths.get(cell.column_letter, 0), len(str(cell.value))))

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width + 2  # Ajustando a largura das colunas para o conteúdo

    # Adicionando aba "Log"
    if log_changes:
        log_file = os.path.join(BASE_DIRECTORY, "log_alteracoes.xlsx")
        df_log = pd.DataFrame(log_changes)
        if os.path.exists(log_file):
            df_existing_log = pd.read_excel(log_file)
            df_log = pd.concat([df_existing_log, df_log], ignore_index=True)
        
        # Criando ou abrindo a aba "Log" na planilha existente
        log_sheet = wb.create_sheet("Log")
        for r_idx, row in df_log.iterrows():
            for c_idx, value in enumerate(row):
                log_sheet.cell(row=r_idx+2, column=c_idx+1, value=value)
        
        # Estilizando a aba "Log"
        log_sheet.insert_rows(1)
        for col in range(1, len(df_log.columns) + 1):
            log_sheet.cell(row=1, column=col).value = df_log.columns[col-1]
            log_sheet.cell(row=1, column=col).font = Font(bold=True)
            log_sheet.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
            log_sheet.cell(row=1, column=col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Ajuste de largura das colunas da aba Log
        column_widths_log = {}
        for row in log_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    column_widths_log[cell.column_letter] = max((column_widths_log.get(cell.column_letter, 0), len(str(cell.value))))

        for col, width in column_widths_log.items():
            log_sheet.column_dimensions[col].width = width + 2  # Ajustando a largura das colunas para o conteúdo

        wb.save(FILE_PATH)
    
    print(f"Planilha atualizada com sucesso: {FILE_PATH}")

# Filtros para e-mails
filters_protestos = ["Aumento em Protesto Detectado!", "Redução de Protestos Detectada!"]
filters_ccf = ["Inclusão de CCF Detectada!"]
filters_processos = ["AUMENTO EM PROCESSOS JUDICIAIS", "REDUÇÃO DE PROCESSOS JUDICIAIS DETECTADA!"]

data_protestos = process_emails(filters_protestos, "Protestos")
data_ccf = process_emails(filters_ccf, "CCF")
data_processos = process_emails(filters_processos, "Processos")

all_data = data_protestos + data_ccf + data_processos

update_excel(all_data)
